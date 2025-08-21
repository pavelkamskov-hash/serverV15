#!/usr/bin/env python3
"""
server.py

This module implements a minimal HTTP server for the extrusion monitoring
dashboard using Python's built‑in libraries.  It provides endpoints
compatible with the required specification without relying on external
packages such as Express.  State is persisted in an SQLite database and
Excel reports are generated with openpyxl.  Sessions are handled via
HTTP cookies stored in memory.
"""

import os
import json
import sqlite3
import threading
import time
import uuid
import urllib.parse
import urllib.request
from http import cookies
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler
from socketserver import ThreadingTCPServer
from urllib.parse import urlparse, parse_qs, unquote

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Determine base directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
PUBLIC_DIR = os.path.join(BASE_DIR, 'public')
SETTINGS_PATH = os.path.join(BASE_DIR, 'config.json')

# -----------------------------------------------------------------------------
# Settings management

def load_settings():
    """Load runtime settings from config.json."""
    try:
        with open(SETTINGS_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def save_settings(cfg):
    """Write settings back to config.json."""
    tmp_path = SETTINGS_PATH + '.tmp'
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, SETTINGS_PATH)


settings = load_settings()

# -----------------------------------------------------------------------------
# Database initialisation

# Ensure the data directory exists
os.makedirs(DATA_DIR, exist_ok=True)

db = sqlite3.connect(os.path.join(DATA_DIR, 'data.db'), check_same_thread=False)
db.row_factory = sqlite3.Row

def init_db():
    cur = db.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS status (
            lineId TEXT PRIMARY KEY,
            isRunning INTEGER DEFAULT 0,
            lastPulseTime INTEGER DEFAULT 0,
            lastPacketTime INTEGER DEFAULT 0
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS status_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lineId TEXT,
            timestamp INTEGER,
            isRunning INTEGER,
            productName TEXT
        )
        """
    )
    # Ensure productName column exists for backwards compatibility.  On
    # older databases created before this column was introduced the
    # column may be missing.  Use PRAGMA to introspect and add it
    # dynamically.
    cur.execute("PRAGMA table_info(status_log)")
    cols = [row[1] for row in cur.fetchall()]
    if 'productName' not in cols:
        cur.execute('ALTER TABLE status_log ADD COLUMN productName TEXT')
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pulses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lineId TEXT,
            pulses INTEGER,
            duration INTEGER,
            timestamp INTEGER
        )
        """
    )
    # Seed 13 lines
    for i in range(1, 14):
        lid = f'line{i}'
        cur.execute('INSERT OR IGNORE INTO status(lineId) VALUES (?)', (lid,))
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS minute_stats (
            lineId TEXT,
            ts INTEGER,
            speed REAL,
            PRIMARY KEY(lineId, ts)
        )
        """
    )
    db.commit()


init_db()

# -----------------------------------------------------------------------------
# Agent implementation (smoothing and state detection)

class Agent:
    """Simple smoothing and state detection engine."""

    def __init__(self, db_conn, config):
        self.db = db_conn
        self.settings = dict(config)
        self.lines = {}
        self.minute_map = {}
        self.lock = threading.Lock()

    def update_settings(self, new_cfg):
        with self.lock:
            self.settings.update(new_cfg or {})

    def ingest(self, line_id, pulses, duration_ms, ts=None):
        ts = int(ts) if ts is not None else int(time.time())
        # convert ms timestamp
        if ts > 1e12:
            ts = ts // 1000
        with self.lock:
            line = self.lines.get(line_id)
            if not line:
                line = {
                    'messages': [],
                    'lastPacket': ts,
                    'lastState': 0,
                    'holdStart': 0,
                    'holdStop': 0,
                    'smoothedSpeed': 0.0,
                }
                self.lines[line_id] = line
            # Append message
            line['messages'].append({'pulses': pulses, 'duration': duration_ms, 'ts': ts})
            line['lastPacket'] = ts
            # Prune messages outside window
            window_sec = int(self.settings.get('windowSec', 60))
            cutoff = ts - window_sec
            line['messages'] = [m for m in line['messages'] if m['ts'] >= cutoff]
            # Compute smoothed speed
            total_pulses = sum(m['pulses'] for m in line['messages'])
            total_duration = sum(m['duration'] for m in line['messages'])
            if total_duration > 0:
                smoothed = (total_pulses / (total_duration / 1000.0)) * 60.0
            else:
                smoothed = 0.0
            line['smoothedSpeed'] = smoothed
            # State detection
            V_START = float(self.settings.get('V_START', 0.5))
            V_STOP = float(self.settings.get('V_STOP', 0.3))
            delay_start = int(self.settings.get('delayStart', 30))
            delay_stop = int(self.settings.get('delayStop', 30))
            new_state = line['lastState']
            if line['lastState'] == 1:
                if smoothed <= V_STOP:
                    if not line['holdStop']:
                        line['holdStop'] = ts
                    if ts - line['holdStop'] >= delay_stop:
                        new_state = 0
                        line['holdStop'] = 0
                        line['holdStart'] = 0
                else:
                    line['holdStop'] = 0
            else:
                if smoothed >= V_START:
                    if not line['holdStart']:
                        line['holdStart'] = ts
                    if ts - line['holdStart'] >= delay_start:
                        new_state = 1
                        line['holdStart'] = 0
                        line['holdStop'] = 0
                else:
                    line['holdStart'] = 0
            state_changed = new_state != line['lastState']
            line['lastState'] = new_state
            # Minute aggregation
            minute_ts = (ts // 60) * 60
            key = (line_id, minute_ts)
            entry = self.minute_map.get(key)
            if not entry:
                entry = {'sum': 0.0, 'count': 0}
                self.minute_map[key] = entry
            entry['sum'] += smoothed
            entry['count'] += 1
            return smoothed, state_changed, new_state

    def flush_minutes(self):
        """Write aggregates older than the current minute to the database."""
        with self.lock:
            now_minute = (int(time.time()) // 60) * 60
            keys = list(self.minute_map.keys())
            for key in keys:
                line_id, minute_ts = key
                if minute_ts < now_minute:
                    entry = self.minute_map.pop(key)
                    avg = entry['sum'] / entry['count'] if entry['count'] > 0 else 0.0
                    cur = self.db.cursor()
                    cur.execute(
                        'INSERT OR REPLACE INTO minute_stats(lineId, ts, speed) VALUES (?,?,?)',
                        (line_id, minute_ts, avg)
                    )
                    cur.close()
            self.db.commit()

    def handle_offline(self, line_id):
        """Reset smoothing for a line if no packets arrived within the timeout."""
        with self.lock:
            line = self.lines.get(line_id)
            if not line:
                return False
            changed = False
            line['messages'] = []
            line['smoothedSpeed'] = 0.0
            line['holdStart'] = 0
            line['holdStop'] = 0
            if line['lastState'] == 1:
                line['lastState'] = 0
                changed = True
            return changed

    def get_smoothed_speed(self, line_id):
        with self.lock:
            line = self.lines.get(line_id)
            return line['smoothedSpeed'] if line else 0.0

    def get_state(self, line_id):
        with self.lock:
            line = self.lines.get(line_id)
            return line['lastState'] if line else 0

    def get_series(self, line_id, hours):
        """Return arrays of ISO timestamps and speeds for the last `hours` hours."""
        now_ts = int(time.time())
        to_minute = (now_ts // 60) * 60
        from_minute = to_minute - int(hours) * 3600
        cur = self.db.cursor()
        cur.execute(
            'SELECT ts, speed FROM minute_stats WHERE lineId=? AND ts>=? AND ts<=? ORDER BY ts ASC',
            (line_id, from_minute, to_minute)
        )
        rows = cur.fetchall()
        cur.close()
        speed_map = {row['ts']: row['speed'] for row in rows}
        labels = []
        data = []
        for t in range(from_minute, to_minute + 60, 60):
            labels.append(time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime(t)))
            data.append(speed_map.get(t, None))
        return {'labels': labels, 'data': data}

    def get_daily_work_idle(self, line_id, days):
        """Compute daily run/down durations for the last `days` days."""
        now_ts = int(time.time())
        to_ts = now_ts
        from_ts = to_ts - days * 86400
        cur = self.db.cursor()
        # Fetch logs in range
        cur.execute(
            'SELECT timestamp, isRunning FROM status_log WHERE lineId=? AND timestamp>=? AND timestamp<=? ORDER BY timestamp ASC',
            (line_id, from_ts, to_ts)
        )
        events = cur.fetchall()
        # Fetch state before range
        cur.execute(
            'SELECT isRunning FROM status_log WHERE lineId=? AND timestamp<? ORDER BY timestamp DESC LIMIT 1',
            (line_id, from_ts)
        )
        row = cur.fetchone()
        cur.close()
        initial_state = row['isRunning'] if row else 0
        labels = []
        work = []
        down = []
        prev_state = initial_state
        # Iterate from oldest day to most recent
        for d in range(days - 1, -1, -1):
            day_start = ((to_ts - d * 86400) // 86400) * 86400
            day_end = day_start + 86400
            # events in this day
            day_events = [e for e in events if day_start <= e['timestamp'] < day_end]
            state = prev_state
            seg_start = day_start
            segments = []
            for ev in day_events:
                ev_state = int(ev['isRunning'])
                if ev_state != state:
                    segments.append({'start': seg_start, 'end': ev['timestamp'], 'state': state})
                    state = ev_state
                    seg_start = ev['timestamp']
            segments.append({'start': seg_start, 'end': day_end, 'state': state})
            prev_state = state
            # merge short segments <60s
            merged = []
            for seg in segments:
                dur = seg['end'] - seg['start']
                if seg['state'] == 1 and dur < 60:
                    # short run -> down
                    if merged and merged[-1]['state'] == 0:
                        merged[-1]['end'] = seg['end']
                    else:
                        merged.append({'start': seg['start'], 'end': seg['end'], 'state': 0})
                elif seg['state'] == 0 and dur < 60:
                    # short stop -> up
                    if merged and merged[-1]['state'] == 1:
                        merged[-1]['end'] = seg['end']
                    else:
                        merged.append({'start': seg['start'], 'end': seg['end'], 'state': 1})
                else:
                    merged.append(dict(seg))
            run_sec = 0
            down_sec = 0
            for seg in merged:
                dur = seg['end'] - seg['start']
                if seg['state'] == 1:
                    run_sec += dur
                else:
                    down_sec += dur
            labels.append(time.strftime('%Y-%m-%d', time.gmtime(day_start)))
            work.append(round(run_sec / 3600.0, 1))
            down.append(round(down_sec / 3600.0, 1))
        return {'labels': labels, 'work': work, 'down': down}


agent = Agent(db, settings)

# -----------------------------------------------------------------------------
# Telegram helper

def send_telegram_message(text):
    """Send a notification via Telegram bot if credentials are set.

    The message is sent asynchronously (fire‑and‑forget).  Errors are
    suppressed to avoid disrupting the main flow.
    """
    token = settings.get('telegramToken')
    chat_id = settings.get('telegramChatId')
    if not token or not chat_id:
        return
    try:
        # Build URL with URL‑encoded parameters
        params = urllib.parse.urlencode({'chat_id': chat_id, 'text': text})
        url = f"https://api.telegram.org/bot{token}/sendMessage?{params}"
        # Fire request without blocking; ignore any returned data
        with urllib.request.urlopen(url, timeout=10) as _:
            pass
    except Exception:
        pass

# -----------------------------------------------------------------------------
# Session management

SESSIONS = {}
SESSION_LOCK = threading.Lock()

def generate_sid():
    return uuid.uuid4().hex

def parse_cookies(header):
    if not header:
        return {}
    c = cookies.SimpleCookie()
    c.load(header)
    return {k: morsel.value for k, morsel in c.items()}

def get_session(handler):
    header = handler.headers.get('Cookie')
    jar = parse_cookies(header)
    sid = jar.get('sid')
    with SESSION_LOCK:
        sess = SESSIONS.get(sid)
        return sid, sess

def set_session_cookie(handler, sid):
    cookie = cookies.SimpleCookie()
    cookie['sid'] = sid
    cookie['sid']['path'] = '/'
    # session cookie (no expiry)
    handler.send_header('Set-Cookie', cookie.output(header=''))

# -----------------------------------------------------------------------------
# HTTP request handler

class Handler(BaseHTTPRequestHandler):
    server_version = 'ExtrusionMonitor/1.0'

    def end_headers(self):
        # Allow CORS for local use if needed
        self.send_header('Access-Control-Allow-Origin', '*')
        super().end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        # Public assets
        if path.startswith('/public/'):
            return self.serve_static(path[len('/public/'):])
        # Health and time endpoints
        if path == '/healthz':
            return self.send_json({'ok': True})
        if path == '/time':
            return self.send_json({'now': int(time.time())})
        # Login page
        if path == '/login':
            return self.serve_static('login.html')
        # Report download proxies
        if path == '/report/last30days':
            # Redirect internally to report_clean
            return self.handle_report_clean()
        # Determine session
        sid, sess = get_session(self)
        # Handle unauthenticated access
        if not sess or sess.get('user') != settings.get('loginUsername'):
            # Only allow access to /login and static
            if path in ('/login',) or path.startswith('/public/'):
                return self.serve_static('login.html')
            # Otherwise redirect to login
            self.send_response(302)
            self.send_header('Location', '/login')
            self.end_headers()
            return
        # Authenticated from here on
        if path == '/':
            return self.serve_static('index.html')
        if path == '/status':
            return self.handle_status()
        if path.startswith('/chartdata/'):
            line_id = path[len('/chartdata/'):]
            return self.handle_chartdata(line_id)
        if path == '/report':
            return self.handle_report(parsed.query)
        if path == '/report_clean':
            return self.handle_report_clean()
        if path == '/settings':
            return self.serve_static('settings.html')
        if path == '/settings/info':
            return self.handle_settings_info(sess)
        # Fallback 404
        self.send_response(404)
        self.end_headers()
        self.wfile.write(b'Not found')

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length) if length > 0 else b''
        content_type = self.headers.get('Content-Type', '')
        data = {}
        if content_type.startswith('application/json'):
            try:
                data = json.loads(body.decode('utf-8') or '{}')
            except Exception:
                data = {}
        elif content_type.startswith('application/x-www-form-urlencoded'):
            qs = parse_qs(body.decode('utf-8'))
            data = {k: v[0] for k, v in qs.items()}
        # Handle login separately as it doesn't require session
        if path == '/login':
            return self.handle_login(data)
        # Session check
        sid, sess = get_session(self)
        if not sess or sess.get('user') != settings.get('loginUsername'):
            self.send_response(401)
            self.end_headers()
            self.wfile.write(b'Unauthorized')
            return
        if path == '/data':
            return self.handle_data(data)
        if path == '/settings/auth':
            return self.handle_settings_auth(data, sid, sess)
        if path == '/settings/save':
            return self.handle_settings_save(data, sid, sess)
        # Unknown POST
        self.send_response(404)
        self.end_headers()
        self.wfile.write(b'Not found')

    # ---------------------------------------------------------------------
    # Static file serving
    def serve_static(self, rel_path):
        # Prevent directory traversal
        safe_path = os.path.normpath(rel_path).lstrip('/\\')
        abs_path = os.path.join(PUBLIC_DIR, safe_path)
        if not abs_path.startswith(PUBLIC_DIR):
            self.send_response(403)
            self.end_headers()
            self.wfile.write(b'Forbidden')
            return
        if not os.path.isfile(abs_path):
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b'Not found')
            return
        ext = os.path.splitext(abs_path)[1].lower()
        mime = 'text/plain'
        if ext == '.html':
            mime = 'text/html; charset=utf-8'
        elif ext == '.css':
            mime = 'text/css'
        elif ext == '.js':
            mime = 'application/javascript'
        elif ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
            mime = f'image/{ext[1:]}'
        elif ext == '.svg':
            mime = 'image/svg+xml'
        try:
            with open(abs_path, 'rb') as f:
                content = f.read()
            self.send_response(200)
            self.send_header('Content-Type', mime)
            self.send_header('Content-Length', str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        except Exception:
            self.send_response(500)
            self.end_headers()
            self.wfile.write(b'Error')

    # ---------------------------------------------------------------------
    # Response helpers
    def send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    # ---------------------------------------------------------------------
    # POST handlers
    def handle_login(self, data):
        username_in = str(data.get('username', ''))
        password_in = str(data.get('password', ''))
        if (username_in == settings.get('loginUsername') and password_in == settings.get('loginPassword')):
            # create session
            sid = generate_sid()
            with SESSION_LOCK:
                SESSIONS[sid] = {'user': username_in, 'settingsAuth': False}
            self.send_response(302)
            set_session_cookie(self, sid)
            self.send_header('Location', '/')
            self.end_headers()
            return
        # Invalid credentials
        self.send_response(401)
        self.end_headers()
        self.wfile.write(b'Unauthorized')

    def handle_data(self, data):
        line_id = str(data.get('lineId', '')).strip()
        pulses = data.get('pulses')
        duration = data.get('duration')
        ts = data.get('ts')
        try:
            p = int(pulses)
            dur = int(duration)
        except Exception:
            return self.send_json({'error': 'invalid payload'}, status=400)
        if not line_id or dur <= 0:
            return self.send_json({'error': 'invalid payload'}, status=400)
        # Insert raw packet
        pkt_ts = int(ts) if ts is not None else int(time.time())
        if pkt_ts > 1e12:
            pkt_ts = pkt_ts // 1000
        cur = db.cursor()
        cur.execute(
            'INSERT INTO pulses(lineId,pulses,duration,timestamp) VALUES (?,?,?,?)',
            (line_id, p, dur, pkt_ts)
        )
        # Ensure line exists
        cur.execute('INSERT OR IGNORE INTO status(lineId) VALUES (?)', (line_id,))
        db.commit()
        # Feed agent
        smoothed, state_changed, new_state = agent.ingest(line_id, p, dur, ts)
        now_sec = int(time.time())
        # Update status table
        sets = ['lastPacketTime=?']
        params = [now_sec]
        if p > 0:
            sets.append('lastPulseTime=?')
            params.append(now_sec)
        if state_changed:
            sets.append('isRunning=?')
            params.append(int(new_state))
        params.append(line_id)
        cur.execute(f'UPDATE status SET {", ".join(sets)} WHERE lineId=?', params)
        if state_changed:
            # Capture current product for this line when logging the event
            product_name = settings.get('products', {}).get(line_id, '')
            cur.execute(
                'INSERT INTO status_log(lineId,timestamp,isRunning,productName) VALUES (?,?,?,?)',
                (line_id, now_sec, int(new_state), product_name)
            )
            # Send Telegram notification (fire and forget)
            display_name = settings.get('lineNames', {}).get(line_id, line_id)
            if int(new_state) == 1:
                action = 'запуск'
            else:
                action = 'остановка'
            msg = f"Линия {display_name} (изделие: {product_name}): {action}"
            send_telegram_message(msg)
        db.commit()
        return self.send_json({'ok': True})

    def handle_settings_auth(self, data, sid, sess):
        password = str(data.get('password', ''))
        settings_pass = settings.get('settingsPassword', '19910509')
        if password == settings_pass:
            with SESSION_LOCK:
                sess['settingsAuth'] = True
            return self.send_json({'ok': True})
        return self.send_json({'error': 'wrong password'}, status=401)

    def handle_settings_save(self, data, sid, sess):
        # Persist new settings from the settings UI.  Since this handler
        # modifies the module‑level ``settings`` object it must declare
        # ``settings`` as global at the beginning of the function.  Python
        # requires the global statement before any use of the variable in
        # the function scope.
        global settings
        if not sess.get('settingsAuth'):
            return self.send_json({'error': 'unauthorized'}, status=401)
        # build new config
        new_cfg = dict(settings)
        def to_num(v, default):
            try:
                n = float(v)
                if n >= 0:
                    return n
            except Exception:
                pass
            return default
        new_cfg['windowSec'] = int(to_num(data.get('windowSec'), settings.get('windowSec', 60)))
        new_cfg['V_START'] = float(to_num(data.get('V_START'), settings.get('V_START', 0.5)))
        new_cfg['V_STOP'] = float(to_num(data.get('V_STOP'), settings.get('V_STOP', 0.3)))
        new_cfg['delayStart'] = int(to_num(data.get('delayStart'), settings.get('delayStart', 30)))
        new_cfg['delayStop'] = int(to_num(data.get('delayStop'), settings.get('delayStop', 30)))
        gh = int(to_num(data.get('graphHours'), settings.get('graphHours', 24)))
        new_cfg['graphHours'] = gh if gh in (24, 48) else settings.get('graphHours', 24)
        new_cfg['offlineTimeout'] = int(to_num(data.get('offlineTimeout'), settings.get('offlineTimeout', 60)))
        if isinstance(data.get('enabledLines'), list):
            new_cfg['enabledLines'] = [str(x) for x in data.get('enabledLines')]
        if isinstance(data.get('lineNames'), dict):
            names = {}
            for i in range(1, 14):
                lid = f'line{i}'
                names[lid] = str(data['lineNames'].get(lid, settings.get('lineNames', {}).get(lid, '')))
            new_cfg['lineNames'] = names
        # Update product names
        if isinstance(data.get('products'), dict):
            prods = {}
            for i in range(1, 14):
                lid = f'line{i}'
                prods[lid] = str(data['products'].get(lid, settings.get('products', {}).get(lid, '')))
            new_cfg['products'] = prods
        # Update Telegram credentials
        if 'telegramToken' in data:
            new_cfg['telegramToken'] = str(data.get('telegramToken') or settings.get('telegramToken', ''))
        if 'telegramChatId' in data:
            new_cfg['telegramChatId'] = str(data.get('telegramChatId') or settings.get('telegramChatId', ''))
        # Preserve credentials and settings password
        new_cfg['loginUsername'] = settings.get('loginUsername')
        new_cfg['loginPassword'] = settings.get('loginPassword')
        new_cfg['settingsPassword'] = settings.get('settingsPassword')
        # Apply and persist
        settings = new_cfg
        save_settings(settings)
        agent.update_settings(settings)
        return self.send_json({'ok': True})

    # ---------------------------------------------------------------------
    # GET handlers
    def handle_status(self):
        # Return status for all 13 lines
        cur = db.cursor()
        cur.execute('SELECT lineId,isRunning,lastPulseTime,lastPacketTime FROM status ORDER BY lineId ASC')
        rows = cur.fetchall()
        cur.close()
        now_sec = int(time.time())
        result = []
        for i in range(1, 14):
            lid = f'line{i}'
            row = next((r for r in rows if r['lineId'] == lid), None)
            if not row:
                row = {'lineId': lid, 'isRunning': 0, 'lastPulseTime': 0, 'lastPacketTime': 0}
            smoothed = agent.get_smoothed_speed(lid)
            state = agent.get_state(lid)
            # Determine label
            offline = True
            if row['lastPacketTime'] and now_sec - row['lastPacketTime'] <= settings.get('offlineTimeout', 60):
                offline = False
            if offline or row['lastPacketTime'] == 0:
                label = 'нет данных'
            else:
                label = 'Работает' if state == 1 else 'Остановлена'
            disp = settings.get('lineNames', {}).get(lid, lid)
            result.append({
                'lineId': lid,
                'displayName': disp,
                'speed': smoothed,
                'isRunning': bool(state),
                'lastPulseTime': row['lastPulseTime'],
                'lastPacketTime': row['lastPacketTime'],
                'stateLabel': label,
                'product': settings.get('products', {}).get(lid, '')
            })
        return self.send_json(result)

    def handle_chartdata(self, line_id):
        hours = int(settings.get('graphHours', 24))
        series = agent.get_series(line_id, hours)
        daily = agent.get_daily_work_idle(line_id, 30)
        line_name = settings.get('lineNames', {}).get(line_id, line_id)
        return self.send_json({'speed': series, 'status': {**daily, 'lineName': line_name}})

    def handle_settings_info(self, sess):
        if not sess.get('settingsAuth'):
            return self.send_json({'error': 'unauthorized'}, status=401)
        return self.send_json(settings)

    def handle_report(self, query):
        if openpyxl is None:
            return self.send_json({'error': 'excel unavailable'}, status=500)
        # parse from and to from query params
        params = parse_qs(query or '')
        def parse_date(val):
            try:
                return int(datetime.datetime.fromisoformat(val).timestamp())
            except Exception:
                return None
        to_ts = None
        from_ts = None
        if 'to' in params:
            to_ts = parse_date(params['to'][0])
        if 'from' in params:
            from_ts = parse_date(params['from'][0])
        now_sec = int(time.time())
        if to_ts is None:
            to_ts = now_sec
        if from_ts is None:
            from_ts = to_ts - 30 * 86400
        # Build workbook
        wb = openpyxl.Workbook()
        summary = wb.active
        summary.title = 'Сводка 30 дней'
        summary.append(['Линия', 'Работа, ч', 'Простой, ч', '% простоя'])
        lines = [f'line{i}' for i in range(1, 14)]
        for lid in lines:
            sheet = wb.create_sheet(title=lid)
            # Add a column for the product (номенклатура)
            sheet.append(['Дата', 'Событие', 'Время', 'Простой, мин', 'Номенклатура'])
            # initial state
            cur = db.cursor()
            cur.execute('SELECT isRunning FROM status_log WHERE lineId=? AND timestamp<? ORDER BY timestamp DESC LIMIT 1', (lid, from_ts))
            row = cur.fetchone()
            initial = row['isRunning'] if row else 0
            # logs
            cur.execute('SELECT timestamp,isRunning,productName FROM status_log WHERE lineId=? AND timestamp>=? AND timestamp<=? ORDER BY timestamp ASC', (lid, from_ts, to_ts))
            events = cur.fetchall()
            cur.close()
            total_run = 0
            total_down = 0
            prev_state = initial
            # Determine last known product before the reporting period
            cur = db.cursor()
            cur.execute('SELECT productName FROM status_log WHERE lineId=? AND timestamp<? ORDER BY timestamp DESC LIMIT 1', (lid, from_ts))
            row = cur.fetchone()
            cur.close()
            prev_product = row['productName'] if row and row['productName'] is not None else settings.get('products', {}).get(lid, '')
            # iterate days
            t = from_ts - (from_ts % 86400)
            while t < to_ts:
                day_start = t
                day_end = min(t + 86400, to_ts)
                day_label = time.strftime('%Y-%m-%d', time.gmtime(day_start))
                day_events = [e for e in events if day_start <= e['timestamp'] < day_end]
                state = prev_state
                seg_start = day_start
                segments = []
                for ev in day_events:
                    ev_state = int(ev['isRunning'])
                    ev_prod = ev['productName']
                    if ev_state != state:
                        # close previous segment, annotate with current product
                        segments.append({'start': seg_start, 'end': ev['timestamp'], 'state': state, 'product': prev_product})
                        # update state and product for next segment
                        state = ev_state
                        seg_start = ev['timestamp']
                        if ev_prod is not None:
                            prev_product = ev_prod
                    elif ev_prod:
                        # product changed without state change
                        segments.append({'start': seg_start, 'end': ev['timestamp'], 'state': state, 'product': prev_product})
                        seg_start = ev['timestamp']
                        prev_product = ev_prod
                # final segment
                segments.append({'start': seg_start, 'end': day_end, 'state': state, 'product': prev_product})
                prev_state = state
                # merge short segments
                merged = []
                for seg in segments:
                    dur = seg['end'] - seg['start']
                    # If a short RUN segment (<60s), merge into previous STOP segment
                    if seg['state'] == 1 and dur < 60:
                        if merged and merged[-1]['state'] == 0:
                            merged[-1]['end'] = seg['end']
                        else:
                            # convert this run into down time
                            merged.append({'start': seg['start'], 'end': seg['end'], 'state': 0, 'product': seg.get('product')})
                    # If a short STOP segment (<60s), merge into previous RUN segment
                    elif seg['state'] == 0 and dur < 60:
                        if merged and merged[-1]['state'] == 1:
                            merged[-1]['end'] = seg['end']
                        else:
                            merged.append({'start': seg['start'], 'end': seg['end'], 'state': 1, 'product': seg.get('product')})
                    else:
                        merged.append(dict(seg))
                # write rows
                for i, seg in enumerate(merged):
                    dur_min = int(round((seg['end'] - seg['start']) / 60))
                    prod = seg.get('product', '')
                    if seg['state'] == 1:
                        total_run += seg['end'] - seg['start']
                        # Работа
                        sheet.append([
                            day_label,
                            'Работа',
                            time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(seg['start'])),
                            str(dur_min),
                            prod
                        ])
                        # Остановка marker
                        sheet.append([
                            day_label,
                            'Остановка',
                            time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(seg['end'])),
                            '',
                            prod
                        ])
                    else:
                        total_down += seg['end'] - seg['start']
                        # Простой
                        sheet.append([
                            day_label,
                            'Простой',
                            time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(seg['start'])),
                            str(dur_min),
                            prod
                        ])
                        if i < len(merged) - 1 and merged[i + 1]['state'] == 1:
                            # Запуск marker at end of stop
                            sheet.append([
                                day_label,
                                'Запуск',
                                time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(seg['end'])),
                                '',
                                prod
                            ])
                t += 86400
            total = total_run + total_down
            pct = f"{(total_down / total * 100):.1f}%" if total > 0 else '0.0%'
            summary.append([lid, f"{total_run / 3600:.1f}", f"{total_down / 3600:.1f}", pct])
        # Remove default sheet if it exists beyond summary
        # Save workbook to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        data = output.getvalue()
        output.close()
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', 'attachment; filename="report.xlsx"')
        self.send_header('Content-Length', str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def handle_report_clean(self):
        if openpyxl is None:
            return self.send_json({'error': 'excel unavailable'}, status=500)
        wb = openpyxl.Workbook()
        summary = wb.active
        summary.title = 'Сводка 30 дней'
        summary.append(['Линия', 'Работа, ч', 'Простой, ч', '% простоя'])
        lines = [f'line{i}' for i in range(1, 14)]
        for lid in lines:
            daily = agent.get_daily_work_idle(lid, 30)
            sheet = wb.create_sheet(title=lid)
            sheet.append(['Дата', 'Работа, ч', 'Простой, ч'])
            run_total = 0.0
            down_total = 0.0
            for lbl, up, down in zip(daily['labels'], daily['work'], daily['down']):
                sheet.append([lbl, up, down])
                run_total += up
                down_total += down
            total = run_total + down_total
            pct = f"{(down_total / total * 100):.1f}%" if total > 0 else '0.0%'
            summary.append([lid, f"{run_total:.1f}", f"{down_total:.1f}", pct])
        # Save workbook
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        data = output.getvalue()
        output.close()
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', 'attachment; filename="report_30days_clean.xlsx"')
        self.send_header('Content-Length', str(len(data)))
        self.end_headers()
        self.wfile.write(data)


# -----------------------------------------------------------------------------
# Background flush and offline watchdog

def background_worker():
    while True:
        try:
            # flush minute aggregates
            agent.flush_minutes()
            # offline detection
            now_sec = int(time.time())
            cur = db.cursor()
            cur.execute('SELECT lineId,isRunning,lastPacketTime FROM status')
            rows = cur.fetchall()
            cur.close()
            for row in rows:
                lid = row['lineId']
                timeout = int(settings.get('offlineTimeout', 60))
                if row['lastPacketTime'] and now_sec - row['lastPacketTime'] > timeout:
                    changed = agent.handle_offline(lid)
                    if changed:
                        c = db.cursor()
                        c.execute('UPDATE status SET isRunning=0 WHERE lineId=?', (lid,))
                        # Log stop event with current product when offline timeout triggers
                        product_name = settings.get('products', {}).get(lid, '')
                        c.execute(
                            'INSERT INTO status_log(lineId,timestamp,isRunning,productName) VALUES (?,?,?,?)',
                            (lid, now_sec, 0, product_name)
                        )
                        # Telegram notification
                        display_name = settings.get('lineNames', {}).get(lid, lid)
                        msg = f"Линия {display_name} (изделие: {product_name}): остановка (нет данных)"
                        send_telegram_message(msg)
                        c.close()
                        db.commit()
        except Exception:
            pass
        time.sleep(15)


# Start background thread
threading.Thread(target=background_worker, daemon=True).start()

# -----------------------------------------------------------------------------
# Launch the server

def run_server(port=3000):
    with ThreadingTCPServer(('0.0.0.0', port), Handler) as httpd:
        print(f'Server running on port {port}')
        httpd.serve_forever()


if __name__ == '__main__':
    run_server(int(os.environ.get('PORT', '3000')))